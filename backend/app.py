"""
Flask Backend API Server for JIRA Connection Testing

This application provides a REST API to test JIRA connections using
Personal Access Token (PAT) authentication.

Author: NDB Date Mover Team
"""

import logging
import os
from typing import Dict, List
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, jsonify, request, send_file, Response
from flask_cors import CORS
from dotenv import load_dotenv
from io import BytesIO, StringIO
import csv
from datetime import datetime

# Optional dependencies for export functionality
REPORTLAB_AVAILABLE = False
OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_TOP
    REPORTLAB_AVAILABLE = True
except ImportError:
    # Set defaults if reportlab not available
    SimpleDocTemplate = None
    Table = None
    TableStyle = None
    Paragraph = None
    colors = None
    landscape = None
    letter = None
    inch = None
    getSampleStyleSheet = None
    ParagraphStyle = None
    TA_TOP = None
    pass  # Will be logged after logger is initialized

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass  # Will be logged in export function

from backend.jira_client import JiraClient, create_jira_client
from backend.config_loader import ConfigLoader, load_config
from backend.date_utils import (
    format_date,
    calculate_week_slip,
    extract_date_history,
    get_week_slip_color,
)
from backend.ai_summarizer import summarize_status_update

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("jira_connection.log"),
        logging.StreamHandler(),
    ],
)

logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-key-change-in-production")

# Enable CORS for frontend - allow all origins for development
CORS(app, resources={
    r"/api/*": {"origins": "*"},
    r"/health": {"origins": "*"}
}, supports_credentials=True)

# Ensure JSON responses are properly formatted
@app.after_request
def after_request(response):
    """Ensure all API responses are JSON."""
    if request.path.startswith('/api/'):
        # Only set JSON content type if not already set and response has content
        if response.content_length:
            # Check if Content-Type is already set (case-insensitive)
            content_type_set = any(
                h.lower() == 'content-type' 
                for h in response.headers.keys()
            )
            if not content_type_set:
                response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response


@app.route("/api/test-connection", methods=["POST"])
def test_connection():
    """
    API endpoint to test JIRA connection.

    This endpoint creates a JiraClient instance and tests the connection
    to JIRA using the credentials from environment variables.

    Returns:
        JSON response with connection test results
    """
    logger.info("Connection test requested")

    try:
        # Create JIRA client from environment variables
        client = create_jira_client()

        if client is None:
            logger.error("Failed to create JIRA client - missing environment variables")
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "JIRA credentials not configured. Please check your .env file.",
                    }
                ),
                400,
            )

        # Use context manager to ensure proper cleanup
        success = False
        result = {}
        try:
            # Test the connection
            success, result = client.test_connection()

            # Optionally get user info if connection is successful
            if success:
                user_info = client.get_user_info()
                if user_info:
                    result["user"] = {
                        "display_name": user_info.get("displayName"),
                        "email": user_info.get("emailAddress"),
                        "account_id": user_info.get("accountId"),
                    }
        finally:
            # Ensure session is always closed
            client.close()

        # Ensure result is always a valid dict
        if not result:
            result = {
                "success": False,
                "message": "Unknown error occurred during connection test"
            }

        # Log the result
        if success:
            logger.info(f"Connection test successful: {result.get('message')}")
        else:
            logger.warning(f"Connection test failed: {result.get('message')}")

        # Return appropriate HTTP status code
        status_code = 200 if success else 400
        return jsonify(result), status_code

    except Exception as e:
        logger.exception("Unexpected error during connection test")
        return (
            jsonify(
                {
                    "success": False,
                    "message": f"Unexpected error: {str(e)}",
                }
            ),
            500,
        )


@app.route("/api/query", methods=["POST"])
def execute_query():
    """
    Execute a JQL query against JIRA and enrich with date history and week slips.
    
    Request body:
        {
            "jql": "project = PROJ AND status = 'In Progress'",
            "max_results": 100,
            "start_at": 0,
            "include_history": true
        }
    
    Returns:
        JSON response with query results enriched with date history
    """
    logger.info("JQL query execution requested")
    
    try:
        # Get request data
        data = request.get_json() or {}
        jql = data.get("jql", "").strip()
        max_results = data.get("max_results", 100)
        start_at = data.get("start_at", 0)
        include_history = data.get("include_history", True)
        
        if not jql:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "JQL query is required",
                    }
                ),
                400,
            )
        
        # Load configuration
        try:
            config_loader = ConfigLoader()
            config = config_loader.load()
            date_fields = config_loader.get_date_fields()
            date_format = config_loader.get_date_format()
        except Exception as e:
            logger.warning(f"Could not load configuration: {str(e)}")
            date_fields = []
            date_format = "mm/dd/yyyy"  # Display format is always mm/dd/yyyy
        
        # Create JIRA client
        client = create_jira_client()
        if client is None:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "JIRA credentials not configured",
                    }
                ),
                400,
            )
        
        try:
            # Execute query
            result = client.execute_jql(jql, max_results, start_at)
            
            if not result.get("success"):
                return jsonify(result), 400
            
            # Get field metadata for display names
            field_metadata = client.get_field_metadata()
            
            # Step 1: Fetch all issues first (already done)
            issues = result.get("issues", [])
            total_issues = len(issues)
            logger.info(f"Fetched {total_issues} issues from JIRA")
            
            # Step 2: Fetch all changelogs in parallel (if history is requested)
            changelog_cache = {}
            if include_history and date_fields:
                # Get list of date fields that track history
                track_history_fields = [
                    field.get("id") for field in date_fields 
                    if field.get("track_history", False)
                ]
                
                if track_history_fields:
                    logger.info(f"Fetching changelogs for {total_issues} issues in parallel...")
                    changelog_cache = fetch_changelogs_parallel(
                        client, issues, track_history_fields, max_workers=10
                    )
                    logger.info(f"Fetched changelogs for {len(changelog_cache)} issues")
            
            # Step 3: Enrich issues with date history and week slips (using cached changelogs)
            logger.info(f"Enriching {total_issues} issues with date history...")
            enriched_issues = []
            for idx, issue in enumerate(issues, 1):
                issue_key = issue.get("key", "UNKNOWN")
                if idx % 10 == 0 or idx == total_issues:
                    logger.info(f"Processing issue {idx}/{total_issues}: {issue_key}")
                
                try:
                    enriched_issue = enrich_issue_with_dates(
                        issue, date_fields, field_metadata, client, 
                        include_history, date_format, changelog_cache
                    )
                    enriched_issues.append(enriched_issue)
                except Exception as e:
                    logger.error(f"Error enriching issue {issue_key}: {str(e)}")
                    # Continue with other issues even if one fails
                    enriched_issues.append(issue)
            
            logger.info(f"Successfully enriched {len(enriched_issues)} issues")
            
            result["issues"] = enriched_issues
            result["field_metadata"] = {
                field_id: {
                    "name": field_data.get("name", field_id),
                    "type": field_data.get("type", "unknown"),
                }
                for field_id, field_data in field_metadata.items()
            }
            
            # Include display_columns from config so frontend knows which columns to show
            try:
                display_cols = config_loader.get_display_columns()
                if display_cols:
                    result["display_columns"] = display_cols
                else:
                    logger.warning("No display_columns found in config, using fallback")
                    result["display_columns"] = None
            except Exception as e:
                logger.warning(f"Error getting display_columns: {str(e)}")
                # Fallback: use all fields if config not available
                result["display_columns"] = None
            
            return jsonify(result), 200
        finally:
            client.close()
            
    except Exception as e:
        logger.exception("Unexpected error during query execution")
        return (
            jsonify(
                {
                    "success": False,
                    "error": f"Unexpected error: {str(e)}",
                }
            ),
            500,
        )


def fetch_changelogs_parallel(
    client: JiraClient,
    issues: List[Dict],
    field_ids: List[str],
    max_workers: int = 10
) -> Dict[str, Dict[str, List[Dict]]]:
    """
    Fetch changelogs for multiple issues in parallel.
    
    This function fetches all changelogs in parallel using ThreadPoolExecutor,
    which significantly reduces the total time compared to sequential fetching.
    
    Args:
        client: JIRA client instance
        issues: List of issue dictionaries
        field_ids: List of field IDs to track
        max_workers: Maximum number of parallel workers (default: 10)
        
    Returns:
        Dict mapping issue_key -> field_id -> list of changes
        Format: {issue_key: {field_id: [change1, change2, ...]}}
    """
    changelog_cache = {}
    
    def fetch_issue_changelog(issue):
        """Fetch changelog for a single issue."""
        issue_key = issue.get("key", "UNKNOWN")
        try:
            # Fetch full changelog for the issue
            changelog = client.get_issue_changelog(issue_key)
            # Organize by field_id for quick lookup
            field_changes = {}
            for change in changelog:
                field_id = change.get("field")
                if field_id in field_ids:
                    if field_id not in field_changes:
                        field_changes[field_id] = []
                    field_changes[field_id].append(change)
            return issue_key, field_changes
        except Exception as e:
            logger.warning(f"Error fetching changelog for {issue_key}: {str(e)}")
            return issue_key, {}
    
    # Fetch changelogs in parallel
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_issue = {
            executor.submit(fetch_issue_changelog, issue): issue 
            for issue in issues
        }
        
        # Collect results as they complete
        completed = 0
        for future in as_completed(future_to_issue):
            completed += 1
            if completed % 10 == 0 or completed == len(issues):
                logger.info(f"Fetched changelogs: {completed}/{len(issues)}")
            
            try:
                issue_key, field_changes = future.result()
                changelog_cache[issue_key] = field_changes
            except Exception as e:
                issue = future_to_issue[future]
                issue_key = issue.get("key", "UNKNOWN")
                logger.error(f"Error processing changelog for {issue_key}: {str(e)}")
                changelog_cache[issue_key] = {}
    
    return changelog_cache


def enrich_issue_with_dates(
    issue: Dict,
    date_fields: List[Dict],
    field_metadata: Dict,
    client: JiraClient,
    include_history: bool,
    date_format: str,
    changelog_cache: Dict[str, Dict[str, List[Dict]]] = None,
) -> Dict:
    """
    Enrich an issue with date history and week slip calculations.
    
    Args:
        issue: JIRA issue data
        date_fields: List of date field configurations
        field_metadata: Field metadata dictionary
        client: JIRA client instance (not used if changelog_cache provided)
        include_history: Whether to fetch and include history
        date_format: Date format string
        changelog_cache: Pre-fetched changelog cache (issue_key -> field_id -> changes)
        
    Returns:
        Dict: Enriched issue data
    """
    issue_key = issue.get("key", "")
    fields = issue.get("fields", {})
    
    # Log available fields for debugging (only for first issue to avoid spam)
    if issue_key and not hasattr(enrich_issue_with_dates, '_logged_fields'):
        logger.debug(f"Available fields in issue {issue_key}: {list(fields.keys())[:20]}...")
        enrich_issue_with_dates._logged_fields = True
    
    # Process each date field that tracks history
    for date_field_config in date_fields:
        field_id = date_field_config.get("id")
        current_value = fields.get(field_id)
        
        if not current_value:
            continue
        
        # Format current date
        formatted_current = format_date(str(current_value), date_format)
        fields[f"{field_id}_formatted"] = formatted_current
        
        if include_history and date_field_config.get("track_history"):
            try:
                # Use cached changelog if available, otherwise fetch on demand
                if changelog_cache and issue_key in changelog_cache:
                    # Get changelog from cache
                    issue_changes = changelog_cache[issue_key]
                    changelog = issue_changes.get(field_id, [])
                else:
                    # Fallback: fetch changelog on demand (for backward compatibility)
                    changelog = client.get_issue_changelog(issue_key, field_id)
                
                date_history = extract_date_history(changelog, field_id)
                
                # Normalize current value for comparison
                from backend.date_display import normalize_date_for_comparison
                current_normalized = normalize_date_for_comparison(str(current_value))
                current_value_str = str(current_value).strip()
                
                # Format historical dates - exclude current date
                formatted_history = []
                seen_dates = set()  # Track seen dates to avoid duplicates
                
                for date_val, _ in date_history:
                    # Normalize both dates for accurate comparison
                    date_normalized = normalize_date_for_comparison(date_val)
                    date_val_str = str(date_val).strip()
                    
                    # Skip if this date matches the current date (multiple comparison methods)
                    should_skip = False
                    
                    # Method 1: Normalized comparison (most reliable)
                    if date_normalized and current_normalized:
                        if date_normalized == current_normalized:
                            should_skip = True
                    
                    # Method 2: String comparison (fallback)
                    if not should_skip:
                        if date_val_str == current_value_str:
                            should_skip = True
                    
                    # Method 3: Compare formatted dates (additional check)
                    # This catches cases where normalization might have failed but formatted dates match
                    if not should_skip:
                        try:
                            formatted_date = format_date(date_val, date_format)
                            formatted_current = format_date(current_value_str, date_format)
                            # Also normalize the formatted dates for comparison (handles 13/Jun/2025 vs 13/Jun/25)
                            formatted_norm1 = normalize_date_for_comparison(formatted_date)
                            formatted_norm2 = normalize_date_for_comparison(formatted_current)
                            if formatted_date == formatted_current or (formatted_norm1 and formatted_norm2 and formatted_norm1 == formatted_norm2):
                                should_skip = True
                        except Exception as e:
                            logger.debug(f"Error in formatted date comparison: {str(e)}")
                            pass  # If formatting fails, continue with other checks
                    
                    if should_skip:
                        continue  # Skip current date
                    
                    # Skip duplicates
                    if date_normalized in seen_dates:
                        continue
                    seen_dates.add(date_normalized or date_val_str)
                    
                    formatted_history.append(format_date(date_val, date_format))
                
                # Reverse chronological order (newest first, oldest last)
                formatted_history.reverse()
                
                fields[f"{field_id}_history"] = formatted_history
                
                # Calculate week slip
                # Use the oldest date from history (first in sorted list) vs current date
                if date_history and len(date_history) > 0:
                    # Count total number of changes (including reverts)
                    # A -> B -> A = 3 changes
                    # Note: date_history includes all dates including current, so changes = len - 1
                    # But we need to count actual transitions, so we count unique dates in history
                    # (excluding current date which is filtered out above)
                    # The number of changes equals the number of historical dates (after filtering current)
                    change_count = len(formatted_history)  # This is the count after filtering out current date
                    fields[f"{field_id}_change_count"] = change_count
                    
                    # Calculate difference from first date to current date
                    original_date = date_history[0][0]  # First date in history (oldest)
                    current_date_str = str(current_value)
                    
                    # Only calculate if we have valid dates
                    if original_date and current_date_str:
                        try:
                            weeks, week_str = calculate_week_slip(original_date, current_date_str)
                            # Ensure week_str is not empty or "N/A" if we have valid dates
                            if not week_str or week_str == "N/A":
                                # Recalculate with better error handling
                                from backend.date_display import calculate_date_difference_display
                                display_str, unit_type = calculate_date_difference_display(original_date, current_date_str)
                                week_str = display_str if display_str != "N/A" else "0 days"
                            
                            fields[f"{field_id}_week_slip"] = {
                                "weeks": weeks,
                                "display": week_str,
                                "color": get_week_slip_color(weeks),
                            }
                        except Exception as e:
                            logger.warning(f"Error calculating week slip for {issue_key}/{field_id}: {str(e)}")
                            fields[f"{field_id}_week_slip"] = {
                                "weeks": 0,
                                "display": "Error",
                                "color": "gray",
                            }
                    else:
                        fields[f"{field_id}_week_slip"] = {
                            "weeks": 0,
                            "display": "N/A",
                            "color": "gray",
                        }
                else:
                    # No history, so no slip and no changes
                    fields[f"{field_id}_change_count"] = 0
                    fields[f"{field_id}_week_slip"] = {
                        "weeks": 0,
                        "display": "N/A",
                        "color": "gray",
                    }
                    
            except Exception as e:
                logger.warning(
                    f"Error enriching date history for {issue_key}/{field_id}: {str(e)}"
                )
                fields[f"{field_id}_history"] = []
                fields[f"{field_id}_week_slip"] = {
                    "weeks": 0,
                    "display": "N/A",
                    "color": "gray",
                }
    
    # Process custom fields that need AI summarization (e.g., status update)
    # Load all custom fields from config to check for AI summarization flags
    try:
        from backend.ai_summarizer import load_ai_prompt
        
        config_loader = ConfigLoader()
        config = config_loader.load()
        custom_fields = config.get("custom_fields", [])
        
        # Load AI prompt from config file
        ai_prompt = load_ai_prompt()
        
        for custom_field in custom_fields:
            field_id = custom_field.get("id")
            # Check if field should be AI summarized (ai_summarize flag)
            if custom_field.get("ai_summarize"):
                field_value = fields.get(field_id)
                logger.debug(f"Processing AI summarization for field {field_id}, value type: {type(field_value)}, value: {str(field_value)[:100] if field_value else 'None'}")
                
                if field_value:
                    # Summarize the field value using the AI prompt
                    try:
                        if isinstance(field_value, str):
                            summarized = summarize_status_update(field_value, prompt=ai_prompt)
                            fields[f"{field_id}_summary"] = summarized
                            fields[f"{field_id}_original"] = field_value  # Keep original for reference
                            # Flag to indicate this field should show AI summary in UI
                            # Defaults to true if ai_summarize is true
                            fields[f"{field_id}_show_ai_summary"] = custom_field.get("show_ai_summary", True)
                            logger.debug(f"AI summary created for {field_id}: {summarized[:100]}")
                        elif isinstance(field_value, dict):
                            # Handle complex field types (e.g., text fields with HTML)
                            # JIRA text fields might be in format: {"value": "text"} or just the text
                            text_value = field_value.get("value") or field_value.get("content") or str(field_value)
                            if text_value and text_value.strip():
                                summarized = summarize_status_update(text_value, prompt=ai_prompt)
                                fields[f"{field_id}_summary"] = summarized
                                fields[f"{field_id}_original"] = text_value
                                fields[f"{field_id}_show_ai_summary"] = custom_field.get("show_ai_summary", True)
                                logger.debug(f"AI summary created for {field_id} (from dict): {summarized[:100]}")
                            else:
                                logger.debug(f"Field {field_id} has empty text value in dict")
                        else:
                            # Try to convert to string
                            text_value = str(field_value).strip()
                            if text_value:
                                summarized = summarize_status_update(text_value, prompt=ai_prompt)
                                fields[f"{field_id}_summary"] = summarized
                                fields[f"{field_id}_original"] = text_value
                                fields[f"{field_id}_show_ai_summary"] = custom_field.get("show_ai_summary", True)
                                logger.debug(f"AI summary created for {field_id} (converted from {type(field_value)}): {summarized[:100]}")
                    except Exception as e:
                        logger.error(f"Error summarizing field {field_id}: {str(e)}", exc_info=True)
                else:
                    logger.debug(f"Field {field_id} has no value (None or empty)")
    except Exception as e:
        logger.warning(f"Error processing AI summarization fields: {str(e)}")
    
    issue["fields"] = fields
    return issue


@app.route("/api/fields", methods=["GET"])
def get_fields():
    """
    Get field metadata from JIRA.
    
    Query parameters:
        field_id: Optional specific field ID to fetch
    
    Returns:
        JSON response with field metadata
    """
    logger.info("Field metadata requested")
    
    try:
        field_id = request.args.get("field_id")
        
        # Create JIRA client
        client = create_jira_client()
        if client is None:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "JIRA credentials not configured",
                    }
                ),
                400,
            )
        
        try:
            # Get field metadata
            fields = client.get_field_metadata(field_id)
            return jsonify({"success": True, "fields": fields}), 200
        finally:
            client.close()
            
    except Exception as e:
        logger.exception("Unexpected error fetching field metadata")
        return (
            jsonify(
                {
                    "success": False,
                    "error": f"Unexpected error: {str(e)}",
                }
            ),
            500,
        )


@app.route("/api/issue/<issue_id>/history", methods=["GET"])
def get_issue_history(issue_id: str):
    """
    Get changelog (change history) for an issue.
    
    Query parameters:
        field_id: Optional specific field ID to filter changes
    
    Returns:
        JSON response with issue change history
    """
    logger.info(f"Changelog requested for issue: {issue_id}")
    
    try:
        field_id = request.args.get("field_id")
        
        # Create JIRA client
        client = create_jira_client()
        if client is None:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "JIRA credentials not configured",
                    }
                ),
                400,
            )
        
        try:
            # Get changelog
            changes = client.get_issue_changelog(issue_id, field_id)
            return jsonify({"success": True, "changes": changes}), 200
        finally:
            client.close()
            
    except Exception as e:
        logger.exception("Unexpected error fetching changelog")
        return (
            jsonify(
                {
                    "success": False,
                    "error": f"Unexpected error: {str(e)}",
                }
            ),
            500,
        )


@app.route("/api/config", methods=["GET"])
def get_config():
    """
    Get the current configuration.
    
    Returns:
        JSON response with configuration data
    """
    try:
        loader = ConfigLoader()
        config = loader.load()
        return jsonify({"success": True, "config": config}), 200
    except FileNotFoundError as e:
        return (
            jsonify(
                {
                    "success": False,
                    "error": str(e),
                }
            ),
            404,
        )
    except Exception as e:
        logger.exception("Unexpected error loading configuration")
        return (
            jsonify(
                {
                    "success": False,
                    "error": f"Unexpected error: {str(e)}",
                }
            ),
            500,
        )


def prepare_export_data(issues: List[Dict], field_metadata: Dict, display_columns: List[str]) -> tuple:
    """
    Prepare data for export with separate date diff columns.
    
    Returns:
        tuple: (headers, rows) where headers includes date diff columns
    """
    if not issues:
        return [], []
    
    # Get date fields from config
    config_loader = ConfigLoader()
    date_fields = config_loader.get_date_fields()
    date_field_ids = [f.get("id") for f in date_fields if f.get("track_history")]
    
    # Build headers - include date diff columns for date fields
    headers = []
    date_diff_columns = {}  # Map field_id to its diff column index
    
    for col in display_columns:
        headers.append(field_metadata.get(col, {}).get("name", col))
        # If this is a date field, add a date diff column after it
        if col in date_field_ids:
            field_name = field_metadata.get(col, {}).get("name", col)
            headers.append(f"{field_name} - Date Difference")
            date_diff_columns[col] = len(headers) - 1
    
    # Build rows
    rows = []
    for issue in issues:
        row = []
        fields = issue.get("fields", {})
        
        for col in display_columns:
            # Get the actual field value
            if col == 'key':
                value = issue.get('key', issue.get('id', ''))
            else:
                value = fields.get(col)
            
            # Format the value for display
            if value is None:
                display_value = ''
            elif isinstance(value, dict):
                display_value = value.get('displayName') or value.get('name') or value.get('value') or str(value)
            elif isinstance(value, list):
                display_value = ', '.join([str(v.get('name') or v.get('displayName') or v) for v in value])
            else:
                display_value = str(value)
            
            # For date fields, use formatted date if available
            if col in date_field_ids and fields.get(f"{col}_formatted"):
                display_value = fields.get(f"{col}_formatted", display_value)
            
            row.append(display_value)
            
            # Add date diff column if this is a date field
            if col in date_diff_columns:
                week_slip = fields.get(f"{col}_week_slip", {})
                if isinstance(week_slip, dict):
                    diff_display = week_slip.get("display", "N/A")
                else:
                    diff_display = "N/A"
                row.append(diff_display)
        
        rows.append(row)
    
    return headers, rows


@app.route("/api/export/<format>", methods=["POST"])
def export_data(format):
    """Export data in the specified format (pdf, csv, excel)."""
    if format == "pdf":
        if not REPORTLAB_AVAILABLE:
            return jsonify({"success": False, "error": "PDF export requires reportlab library. Install with: pip install reportlab"}), 500
        return export_pdf()
    elif format == "csv":
        return export_csv()
    elif format == "excel":
        if not OPENPYXL_AVAILABLE:
            return jsonify({"success": False, "error": "Excel export requires openpyxl library. Install with: pip install openpyxl"}), 500
        return export_excel()
    else:
        return jsonify({"success": False, "error": f"Unsupported format: {format}"}), 400


def prepare_pdf_data(issues: List[Dict], field_metadata: Dict, display_columns: List[str]) -> tuple:
    """
    Prepare data for PDF export matching UI display (no separate date diff columns).
    
    Returns:
        tuple: (headers, rows) matching what's shown in the UI
    """
    if not issues:
        return [], []
    
    # Get date fields from config
    config_loader = ConfigLoader()
    date_fields = config_loader.get_date_fields()
    date_field_ids = [f.get("id") for f in date_fields if f.get("track_history")]
    
    # Build headers - just the display columns, no separate date diff columns
    headers = []
    for col in display_columns:
        headers.append(field_metadata.get(col, {}).get("name", col))
    
    # Build rows matching UI display
    rows = []
    for issue in issues:
        row = []
        fields = issue.get("fields", {})
        
        for col in display_columns:
            # Get the actual field value
            if col == 'key':
                value = issue.get('key', issue.get('id', ''))
                display_value = str(value) if value else ''
            else:
                value = fields.get(col)
                
                # Check if this is a date field with formatted display
                if col in date_field_ids and fields.get(f"{col}_formatted"):
                    # For date fields, show current date with history info inline
                    current = fields.get(f"{col}_formatted", "")
                    history = fields.get(f"{col}_history", [])
                    change_count = fields.get(f"{col}_change_count", 0)
                    week_slip = fields.get(f"{col}_week_slip", {})
                    
                    # Build display like UI: current date, then history, then metrics
                    # Truncate for PDF to avoid cell overflow
                    display_parts = [current] if current else []
                    
                    if history:
                        # Limit history to first 3 dates for PDF
                        history_display = history[:3] if len(history) > 3 else history
                        history_str = " | ".join([f"~~{h}~~" for h in history_display])
                        if len(history) > 3:
                            history_str += f" (+{len(history) - 3} more)"
                        display_parts.append(history_str)
                    
                    if change_count > 0:
                        display_parts.append(f"Chg:{change_count}")
                    
                    if isinstance(week_slip, dict) and week_slip.get("display"):
                        display_parts.append(week_slip.get("display", ""))
                    
                    display_value = " | ".join(display_parts)
                    # Truncate total length to 100 chars for PDF
                    if len(display_value) > 100:
                        display_value = display_value[:97] + "..."
                elif value is None:
                    display_value = ''
                elif isinstance(value, dict):
                    display_value = value.get('displayName') or value.get('name') or value.get('value') or str(value)
                elif isinstance(value, list):
                    display_value = ', '.join([str(v.get('name') or v.get('displayName') or v) for v in value])
                else:
                    display_value = str(value)
                
                # Truncate long values for PDF (max 80 chars per cell)
                if len(display_value) > 80:
                    display_value = display_value[:77] + "..."
            
            row.append(display_value)
        
        rows.append(row)
    
    return headers, rows


def export_pdf():
    """Export data as PDF in landscape mode matching UI display."""
    try:
        data = request.get_json() or {}
        issues = data.get("issues", [])
        field_metadata = data.get("field_metadata", {})
        display_columns = data.get("display_columns", [])
        
        if not issues:
            return jsonify({"success": False, "error": "No data to export"}), 400
        
        headers, rows = prepare_pdf_data(issues, field_metadata, display_columns)
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                               rightMargin=0.3*inch, leftMargin=0.3*inch,
                               topMargin=0.3*inch, bottomMargin=0.3*inch)
        
        # Create styles for text wrapping
        styles = getSampleStyleSheet()
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=6,
            leading=7,
            alignment=TA_LEFT,
            wordWrap='CJK'  # Enable word wrapping
        )
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=7,
            leading=8,
            alignment=TA_LEFT,
            fontName='Helvetica-Bold',
            wordWrap='CJK'
        )
        
        # Convert table data to Paragraphs for proper text wrapping
        def wrap_text(text, style):
            """Convert text to Paragraph for wrapping."""
            if not text:
                return Paragraph('', style)
            try:
                # Replace special characters that might break PDF rendering
                text = str(text).replace('~~', '')  # Remove strike-through markers
                # Escape HTML/XML special characters for Paragraph
                text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                # Replace quotes that might cause issues
                text = text.replace('"', '&quot;').replace("'", '&#39;')
                return Paragraph(text, style)
            except Exception as e:
                logger.warning(f"Error wrapping text for PDF: {str(e)}, using empty paragraph")
                # Always return a Paragraph object, never a string
                try:
                    return Paragraph(str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')[:500], style)
                except:
                    return Paragraph('', style)
        
        # Build table data with wrapped text
        wrapped_table_data = []
        
        # Headers row
        wrapped_headers = [wrap_text(h, header_style) for h in headers]
        wrapped_table_data.append(wrapped_headers)
        
        # Data rows
        for row in rows:
            wrapped_row = [wrap_text(cell, normal_style) for cell in row]
            wrapped_table_data.append(wrapped_row)
        
        # Create table with column widths (auto-size based on content)
        col_widths = []
        if headers:
            # Estimate width: landscape letter is ~792 points wide, minus margins
            available_width = landscape(letter)[0] - 0.6*inch
            num_cols = len(headers)
            base_width = available_width / num_cols
            col_widths = [base_width] * num_cols
        
        table = Table(wrapped_table_data, colWidths=col_widths if col_widths else None, repeatRows=1)
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        table.setStyle(style)
        
        # Build PDF
        elements = [table]
        doc.build(elements)
        
        buffer.seek(0)
        filename = f"jira_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.exception("Error generating PDF export")
        return jsonify({"success": False, "error": str(e)}), 500


def export_csv():
    """Export data as CSV."""
    try:
        data = request.get_json() or {}
        issues = data.get("issues", [])
        field_metadata = data.get("field_metadata", {})
        display_columns = data.get("display_columns", [])
        
        if not issues:
            return jsonify({"success": False, "error": "No data to export"}), 400
        
        headers, rows = prepare_export_data(issues, field_metadata, display_columns)
        
        # Create CSV in memory (text mode, not binary)
        text_buffer = StringIO()
        writer = csv.writer(text_buffer)
        writer.writerow(headers)
        writer.writerows(rows)
        
        # Get the string content and encode to bytes
        csv_content = text_buffer.getvalue()
        filename = f"jira_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return Response(
            csv_content.encode('utf-8'),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.exception("Error generating CSV export")
        return jsonify({"success": False, "error": str(e)}), 500


def export_excel():
    """Export data as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({"success": False, "error": "openpyxl library not installed"}), 500
    
    try:
        data = request.get_json() or {}
        issues = data.get("issues", [])
        field_metadata = data.get("field_metadata", {})
        display_columns = data.get("display_columns", [])
        
        if not issues:
            return jsonify({"success": False, "error": "No data to export"}), 400
        
        headers, rows = prepare_export_data(issues, field_metadata, display_columns)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "JIRA Export"
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Add rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"jira_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.exception("Error generating Excel export")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/health")
def health():
    """
    Health check endpoint.

    Returns:
        JSON response indicating service health
    """
    return jsonify({"status": "healthy", "service": "jira-connection-tester"}), 200


if __name__ == "__main__":
    # Check if required environment variables are set
    jira_url = os.getenv("JIRA_URL")
    jira_token = os.getenv("JIRA_PAT_TOKEN")

    if not jira_url or not jira_token:
        logger.warning(
            "JIRA_URL or JIRA_PAT_TOKEN not set in environment. "
            "Please configure your .env file."
        )

    # Run the Flask app
    # Using unusual port 8473 to avoid conflicts with common services
    backend_port = int(os.getenv("BACKEND_PORT", "8473"))
    logger.info(f"Starting Flask application on port {backend_port}...")
    app.run(debug=True, host="0.0.0.0", port=backend_port)

